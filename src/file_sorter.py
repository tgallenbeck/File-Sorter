
import glob
from json import load
import os
import pandas as pd
from openpyxl import load_workbook



def sort_files() -> list | str:
    # Setting initial working directory
    directory = os.getenv('DIRECTORY', 'Could not find any directory in the .env file')
    # Changing to the initial working directory
    os.chdir(directory)
    
    # Grab each strategy file from the initial working directory according to the time they were made or downloaded
    '''
    IMPORTANT!! How you download each of the files from your back test matters because of the pre-build structure of the spreadsheet. I've used a quick method of simply sorting to ensure that this works correctly. If you are trading all 28 pairs, then you should export your results first by performance summary and then by list of trades. This should also be in the same order that the spreadsheet has the pairs ordered!
    '''
    security_data = sorted(glob.glob('*.csv'), key=os.path.getmtime)

    # Return the variable storing the files
    return security_data, directory


def sort_list_of_trades(security_data, directory) -> list | list:
    # Split the the files into performance summary and list of trades files. Since they were pre-sorted, the files are split from the first(0) and the second(1) each with a step of 2
    performance_summary = [security_data[0::2]]
    list_of_trades = [security_data[1::2]]

    # Loop to open each list of trades file and sort the trades in ascending order
    for files in list_of_trades:
        for f in files:
            # Read in the files
            df = pd.read_csv(f'{directory}\\{f}', sep=',')
            # Rename 'Trade #' to something that causes less problems
            df.rename(columns = {'Unnamed: 0': 'Trade'}, inplace=True)
            df.rename(columns = {'Trade #': 'Trade'}, inplace = True)
            # Sort the values in ascending order
            sorted_df = df.sort_values(by=['Trade'], ascending=True)
            # Save the CSV
            sorted_df.to_csv(f'{f}', index=False)

    # Return both the sorted list and the list of files containing the performance summary data
    return performance_summary, list_of_trades
    
    
def insert_trade_data_summary(perf_summary, directory, strat_or_indi_backtest_file):

    book = load_workbook(strat_or_indi_backtest_file, keep_vba=True)

    writer = pd.ExcelWriter(strat_or_indi_backtest_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    n = 1
    while n <= 28:
        for files in perf_summary:
            for f in files:
                t_df = pd.read_csv(f'{directory}\\{f}')
                t_df.dropna()

                t_df.to_excel(writer, sheet_name=f'{n}', header=None, index=False, startcol=1, startrow=10)
                
                n = n+1
    writer.close()


def insert_trade_data_list_of_trades(list_of_trades, directory, strat_or_indi_backtest_file):

    book = load_workbook(strat_or_indi_backtest_file, keep_vba=True)

    writer = pd.ExcelWriter(strat_or_indi_backtest_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    n = 1
    while n <= 28:
        for files in list_of_trades:
            for f in files:
                t_df = pd.read_csv(f'{directory}\\{f}')
                t_df.dropna()

                t_df.to_excel(writer, sheet_name=f'{n}', header=None, index=False, startcol=1, startrow=45)
                
                n = n+1
    writer.close()


if __name__ == '__main__':
    # 1. Get a list of all the files for the back test
    security_data, directory = sort_files()

    # 2. Sort each file that contains "List_of_Trades" in the filename from smallest to largest
    perf_summary, list_of_trades = sort_list_of_trades(security_data, directory)

    # 3. Import each data set into "strat_or_indi_backtest_file"
    #   - Each pair of files (performance summary / list of trades) will need to be input on the following:
    #       - Tab Label 1: Performance summary data --> Cell B10
    #       - Tab Label 1: List of Trades --> Cell B45
    #       - Switch to Tab Label 2:
    #       - Tab Label 2: Performance summary --> Cell B10
    #       - Tab Label 2: List of Trades --> Cell B45
    #       - Etc...
    strat_or_indi_backtest_file = input(r"Enter the path of the strategy spreadsheet you are testing: ")
    insert_trade_data_summary(perf_summary, directory, strat_or_indi_backtest_file)
    insert_trade_data_list_of_trades(list_of_trades, directory, strat_or_indi_backtest_file)